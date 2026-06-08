import os

FILE_PATH = "src/phase1_structure.py"

NEW_FUNCTION = """
def build_dhims_summary_sheet(wb, config):
    from openpyxl.worksheet.datavalidation import DataValidation
    ws = wb.create_sheet("DHIMS Summary")
    ws.sheet_properties.tabColor = "008000"

    ws.merge_cells("A1:L1")
    c = ws.cell(row=1, column=1, value="LHIMS DHIMS Reports - Statement Of In-Patient (Age/Gender Summary)")
    c.font = HEADER_FONT
    c.alignment = CENTER

    # Filters
    ws.cell(row=2, column=1, value="Select Ward:").font = BOLD_FONT
    ward_codes = ",".join([w.code for w in config.WARDS])
    dv_ward = DataValidation(type="list", formula1=f'"{ward_codes},All Wards"', allow_blank=True)
    ws.add_data_validation(dv_ward)
    dv_ward.add("B2")
    ws.cell(row=2, column=2, value="All Wards")
    ws.cell(row=2, column=2).border = THIN_BORDER
    ws.cell(row=2, column=2).fill = LIGHT_YELLOW_FILL

    ws.cell(row=3, column=1, value="Select Month:").font = BOLD_FONT
    dv_month = DataValidation(type="list", formula1='"1,2,3,4,5,6,7,8,9,10,11,12,All Months"', allow_blank=True)
    ws.add_data_validation(dv_month)
    dv_month.add("B3")
    ws.cell(row=3, column=2, value="All Months")
    ws.cell(row=3, column=2).border = THIN_BORDER
    ws.cell(row=3, column=2).fill = LIGHT_YELLOW_FILL
    
    # Headers
    ws.merge_cells("C5:F5")
    c = ws.cell(row=5, column=3, value="INSURED PATIENTS")
    c.font = HEADER_FONT_WHITE
    c.fill = HEADER_FILL
    c.alignment = CENTER
    
    ws.merge_cells("G5:J5")
    c = ws.cell(row=5, column=7, value="NON-INSURED PATIENTS")
    c.font = HEADER_FONT_WHITE
    c.fill = HEADER_FILL
    c.alignment = CENTER
    
    ws.merge_cells("C6:D6")
    ws.cell(row=6, column=3, value="Admission").alignment = CENTER
    ws.merge_cells("E6:F6")
    ws.cell(row=6, column=5, value="Death").alignment = CENTER
    ws.merge_cells("G6:H6")
    ws.cell(row=6, column=7, value="Admission").alignment = CENTER
    ws.merge_cells("I6:J6")
    ws.cell(row=6, column=9, value="Death").alignment = CENTER
    ws.merge_cells("K6:L6")
    ws.cell(row=6, column=11, value="Total").alignment = CENTER
    
    headers = [
        "Sr.No.", "Age Groups(Yrs.)",
        "Male", "Female", "Male", "Female",
        "Male", "Female", "Male", "Female",
        "Male", "Female"
    ]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=7, column=col, value=h)
        c.font = BOLD_FONT
        c.alignment = CENTER
        c.border = THIN_BORDER
        c.fill = LIGHT_BLUE_FILL
        
    for col in range(1, 13):
        ws.cell(row=6, column=col).border = THIN_BORDER
        ws.cell(row=6, column=col).font = BOLD_FONT
        
    def _dhims_formula(table, age_units, age_mins, age_maxs, sex, nhis):
        parts = []
        for i in range(len(age_units)):
            conds = []
            conds.append(f'{table}[WardCode], IF($B$2="All Wards", "?*", $B$2)')
            conds.append(f'{table}[Month], IF($B$3="All Months", ">=0", $B$3)')
            conds.append(f'{table}[Sex], "{sex}"')
            conds.append(f'{table}[NHIS], "{nhis}"')
            conds.append(f'{table}[AgeUnit], "{age_units[i]}"')
            if age_mins[i] is not None:
                conds.append(f'{table}[Age], ">={age_mins[i]}"')
            if age_maxs[i] is not None:
                conds.append(f'{table}[Age], "<={age_maxs[i]}"')
            parts.append(f'COUNTIFS({ ", ".join(conds) })')
        return "=IFERROR(" + " + ".join(parts) + ", 0)"

    age_groups = [
        ("0-28 Days", ["Days"], [None], [28]),
        ("1-11 Months", ["Days", "Months"], [29, None], [None, 11]),
        ("1-4", ["Months", "Years"], [12, 1], [None, 4]),
        ("5-9", ["Years"], [5], [9]),
        ("10-14", ["Years"], [10], [14]),
        ("15-17", ["Years"], [15], [17]),
        ("18-19", ["Years"], [18], [19]),
        ("20-34", ["Years"], [20], [34]),
        ("35-49", ["Years"], [35], [49]),
        ("50-59", ["Years"], [50], [59]),
        ("60-69", ["Years"], [60], [69]),
        ("70 & Above", ["Years"], [70], [None])
    ]
    
    current_row = 8
    for idx, (label, units, mins, maxs) in enumerate(age_groups, 1):
        ws.cell(row=current_row, column=1, value=idx).alignment = CENTER
        ws.cell(row=current_row, column=2, value=label).font = BOLD_FONT
        ws.cell(row=current_row, column=3, value=_dhims_formula("tblAdmissions", units, mins, maxs, "M", "Insured"))
        ws.cell(row=current_row, column=4, value=_dhims_formula("tblAdmissions", units, mins, maxs, "F", "Insured"))
        ws.cell(row=current_row, column=5, value=_dhims_formula("tblDeaths", units, mins, maxs, "M", "Insured"))
        ws.cell(row=current_row, column=6, value=_dhims_formula("tblDeaths", units, mins, maxs, "F", "Insured"))
        ws.cell(row=current_row, column=7, value=_dhims_formula("tblAdmissions", units, mins, maxs, "M", "Non-Insured"))
        ws.cell(row=current_row, column=8, value=_dhims_formula("tblAdmissions", units, mins, maxs, "F", "Non-Insured"))
        ws.cell(row=current_row, column=9, value=_dhims_formula("tblDeaths", units, mins, maxs, "M", "Non-Insured"))
        ws.cell(row=current_row, column=10, value=_dhims_formula("tblDeaths", units, mins, maxs, "F", "Non-Insured"))
        
        r = current_row
        ws.cell(row=r, column=11, value=f'=C{r}+E{r}+G{r}+I{r}')
        ws.cell(row=r, column=12, value=f'=D{r}+F{r}+H{r}+J{r}')
        
        for c in range(1, 13):
            ws.cell(row=r, column=c).border = THIN_BORDER
        
        current_row += 1
        
    ws.cell(row=current_row, column=1, value="")
    ws.cell(row=current_row, column=2, value="TOTAL").font = BOLD_FONT
    for c in range(3, 13):
        from openpyxl.utils import get_column_letter
        col_let = get_column_letter(c)
        ws.cell(row=current_row, column=c, value=f'=SUM({col_let}8:{col_let}{current_row-1})').font = BOLD_FONT
        ws.cell(row=current_row, column=c).fill = TOTAL_FILL
        ws.cell(row=current_row, column=c).border = THIN_BORDER
    ws.cell(row=current_row, column=1).border = THIN_BORDER
    ws.cell(row=current_row, column=2).border = THIN_BORDER
    
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 18
    for i in range(3, 13):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = 10

"""

def modify_file():
    with open(FILE_PATH, 'r', encoding='utf-8') as f:
        content = f.read()

    if "build_dhims_summary_sheet" in content:
        print("Function already exists.")
        return

    main_func_idx = content.find('def build_structure')
    if main_func_idx == -1:
        print("Marker not found.")
        return

    # Insert function definition
    content = content[:main_func_idx] + NEW_FUNCTION + "\\n" + content[main_func_idx:]

    # Insert call in main
    main_marker = "    # 8. Statement of Inpatient\\n    build_statement_of_inpatient_sheet(wb, config)"
    new_call = main_marker + "\\n\\n    # 8b. DHIMS Summary\\n    build_dhims_summary_sheet(wb, config)"
    
    if main_marker in content:
        content = content.replace(main_marker, new_call)
    else:
        print("Main marker not found.")
        return

    with open(FILE_PATH, 'w', encoding='utf-8') as f:
        f.write(content)
        
    print("Injection successful.")

if __name__ == "__main__":
    modify_file()
