"""
Phase 1: Build workbook structure using openpyxl
Creates all sheets, Excel Tables, formatting, formulas, and data validation.
"""
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, numbers
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from config import WorkbookConfig

# ── Style constants ──────────────────────────────────────────────────────────

HEADER_FONT = Font(name="Calibri", bold=True, size=14)
SUBHEADER_FONT = Font(name="Calibri", bold=True, size=12)
LABEL_FONT = Font(name="Calibri", bold=True, size=10)
NORMAL_FONT = Font(name="Calibri", size=10)
BOLD_FONT = Font(name="Calibri", bold=True, size=10)

CENTER = Alignment(horizontal="center", vertical="center")
CENTER_WRAP = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center")

THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT_WHITE = Font(name="Calibri", bold=True, size=10, color="FFFFFF")
LIGHT_BLUE_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
LIGHT_GREEN_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
LIGHT_YELLOW_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
GRAY_FILL = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
TOTAL_FILL = PatternFill(start_color="B4C6E7", end_color="B4C6E7", fill_type="solid")

TABLE_STYLE = TableStyleInfo(
    name="TableStyleLight9",
    showFirstColumn=False, showLastColumn=False,
    showRowStripes=True, showColumnStripes=False,
)


def _apply_border_range(ws, min_row, max_row, min_col, max_col):
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).border = THIN_BORDER


# ═══════════════════════════════════════════════════════════════════════════════
# CONTROL SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def build_control_sheet(wb: Workbook, config: WorkbookConfig):
    ws = wb.active
    ws.title = "Control"
    ws.sheet_properties.tabColor = "1F4E79"

    # Header
    ws.merge_cells("A1:H1")
    c = ws.cell(row=1, column=1, value="GHANA HEALTH SERVICE")
    c.font = HEADER_FONT
    c.alignment = CENTER

    ws.merge_cells("A2:H2")
    c = ws.cell(row=2, column=1, value="BED UTILIZATION MANAGEMENT SYSTEM")
    c.font = SUBHEADER_FONT
    c.alignment = CENTER

    ws.merge_cells("A3:H3")
    c = ws.cell(row=3, column=1, value=config.hospital_name)
    c.font = Font(name="Calibri", bold=True, size=11)
    c.alignment = CENTER

    # Year and Hospital info (compact, side by side)
    ws.cell(row=5, column=1, value="Year:").font = BOLD_FONT
    ws.cell(row=5, column=2, value=config.year).font = BOLD_FONT

    ws.cell(row=5, column=4, value="Hospital:").font = BOLD_FONT
    ws.merge_cells("E5:H5")
    ws.cell(row=5, column=5, value=config.hospital_name).font = NORMAL_FONT

    # Hidden preference storage (for VBA access)
    ws.cell(row=5, column=10, value=config.preferences.show_emergency_total_remaining)  # J5
    ws.cell(row=6, column=10, value=config.preferences.subtract_deaths_under_24hrs_from_admissions)  # J6
    ws.column_dimensions["J"].hidden = True

    # Section header for buttons
    ws.merge_cells("A7:H7")
    c = ws.cell(row=7, column=1, value="DATA ENTRY & REPORTS")
    c.font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
    c.alignment = CENTER

    # Button placeholders (actual buttons added by phase2 via win32com)
    buttons = [
        (9,  "[ DAILY BED ENTRY ]",        "Enter daily admissions, discharges, deaths, transfers for each ward"),
        (11, "[ RECORD ADMISSION ]",       "Record individual patient admission details (for age/gender reports)"),
        (13, "[ RECORD DEATH ]",           "Record individual death details (for deaths report)"),
        (15, "[ RECORD AGES GROUP ]",      "Quick entry for age group admissions (bulk mode)"),
        (17, "[ REFRESH REPORTS ]",        "Update Deaths Report and COD Summary sheets"),
        (19, "[ MANAGE WARDS ]",           "Add, edit, or delete ward configurations"),
        (21, "[ EXPORT WARD CONFIG ]",     "Save ward configuration to JSON file for rebuilding"),
        (23, "[ EXPORT YEAR-END ]",        "Export carry-forward data for next year"),
        (25, "[ HOSPITAL PREFERENCES ]",   "Configure hospital preferences with user-friendly form"),
    ]
    rebuild_button = [
        (27, "[ REBUILD WORKBOOK ]",       "Automatically rebuild workbook with new preferences (creates backup)"),
    ]
    diagnostic_buttons = [
        (29, "[ IMPORT OLD WORKBOOK ]",    "Import data from previous workbook format"),
        (31, "[ RECALCULATE ALL DATA ]",   "Recalculate all remaining values"),
        (33, "[ VERIFY CALCULATIONS ]",    "Verify calculation accuracy"),
        (35, "[ FIX DATE FORMATS ]",       "Fix date formatting issues in all data tables"),
    ]
    for row, label, desc in buttons:
        ws.merge_cells(f"A{row}:C{row}")
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(name="Calibri", bold=True, size=11, color="1F4E79")
        c.alignment = CENTER
        c.fill = LIGHT_BLUE_FILL
        c.border = THIN_BORDER
        ws.cell(row=row, column=4, value=desc).font = NORMAL_FONT

    # Rebuild button (special warning color)
    for row, label, desc in rebuild_button:
        ws.merge_cells(f"A{row}:C{row}")
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(name="Calibri", bold=True, size=11, color="FFFFFF")  # White text
        c.alignment = CENTER
        c.fill = PatternFill(start_color="FF6600", end_color="FF6600", fill_type="solid")  # Orange
        c.border = THIN_BORDER
        ws.cell(row=row, column=4, value=desc).font = NORMAL_FONT

    for row, label, desc in diagnostic_buttons:
        ws.merge_cells(f"A{row}:C{row}")
        c = ws.cell(row=row, column=1, value=label)
        c.font = Font(name="Calibri", bold=True, size=10, color="808080")
        c.alignment = CENTER
        c.fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")
        c.border = THIN_BORDER
        ws.cell(row=row, column=4, value=desc).font = NORMAL_FONT

    # ── Ward Configuration Table ─────────────────────────────────────────
    config_start = 37  # Row for section header (after all buttons)

    # Section header for ward configuration
    ws.merge_cells(f"A{config_start}:F{config_start}")
    c = ws.cell(row=config_start, column=1, value="WARD CONFIGURATION")
    c.font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
    c.alignment = CENTER
    c.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    headers = ["WardCode", "WardName", "BedComplement", "PrevYearRemaining", "IsEmergency", "DisplayOrder"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=config_start + 1, column=col, value=h)
        c.font = HEADER_FONT_WHITE
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = THIN_BORDER

    for i, ward in enumerate(config.WARDS):
        r = config_start + 2 + i
        vals = [ward.code, ward.name, ward.bed_complement,
                ward.prev_year_remaining, ward.is_emergency, ward.display_order]
        for col, val in enumerate(vals, 1):
            c = ws.cell(row=r, column=col, value=val)
            c.font = NORMAL_FONT
            c.alignment = CENTER
            c.border = THIN_BORDER

    # Create the table
    end_row = config_start + 1 + len(config.WARDS)
    tab_ref = f"A{config_start + 1}:F{end_row}"
    tbl = Table(displayName="tblWardConfig", ref=tab_ref)
    tbl.tableStyleInfo = TABLE_STYLE
    ws.add_table(tbl)

    # ── Preferences Configuration Table ──────────────────────────────────
    prefs_start = end_row + 2  # end_row from tblWardConfig

    # Section header
    ws.merge_cells(f"A{prefs_start}:F{prefs_start}")
    c = ws.cell(row=prefs_start, column=1, value="HOSPITAL PREFERENCES CONFIGURATION")
    c.font = Font(name="Calibri", bold=True, size=12, color="1F4E79")
    c.alignment = CENTER
    c.fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

    # Headers
    pref_headers = ["PreferenceKey", "PreferenceValue", "Description"]
    for col, h in enumerate(pref_headers, 1):
        c = ws.cell(row=prefs_start + 1, column=col, value=h)
        c.font = HEADER_FONT_WHITE
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = THIN_BORDER

    # Data rows
    pref_rows = [
        ("show_emergency_total_remaining",
         config.preferences.show_emergency_total_remaining,
         "Show 'Emergency Total Remaining' row in Monthly Summary"),
        ("subtract_deaths_under_24hrs_from_admissions",
         config.preferences.subtract_deaths_under_24hrs_from_admissions,
         "Subtract deaths under 24hrs from monthly admission totals")
    ]

    for idx, (key, value, desc) in enumerate(pref_rows, prefs_start + 2):
        ws.cell(row=idx, column=1, value=key).font = NORMAL_FONT
        ws.cell(row=idx, column=2, value=value).font = NORMAL_FONT  # Boolean
        ws.cell(row=idx, column=3, value=desc).font = NORMAL_FONT
        for col in range(1, 4):
            ws.cell(row=idx, column=col).alignment = CENTER
            ws.cell(row=idx, column=col).border = THIN_BORDER

    # Create table
    prefs_end = prefs_start + 1 + len(pref_rows)
    prefs_tab_ref = f"A{prefs_start + 1}:C{prefs_end}"
    prefs_tbl = Table(displayName="tblPreferences", ref=prefs_tab_ref)
    prefs_tbl.tableStyleInfo = TABLE_STYLE
    ws.add_table(prefs_tbl)

    # ── Month Lookup Table ───────────────────────────────────────────────
    month_start = prefs_end + 2  # CHANGED from: month_start = end_row + 2
    ws.cell(row=month_start, column=1, value="MONTH LOOKUP").font = SUBHEADER_FONT

    month_headers = ["MonthNum", "MonthName", "DaysInMonth"]
    for col, h in enumerate(month_headers, 1):
        c = ws.cell(row=month_start + 1, column=col, value=h)
        c.font = HEADER_FONT_WHITE
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = THIN_BORDER

    for m in range(1, 13):
        r = month_start + 1 + m
        ws.cell(row=r, column=1, value=m).font = NORMAL_FONT
        ws.cell(row=r, column=2, value=config.MONTH_NAMES[m - 1]).font = NORMAL_FONT
        ws.cell(row=r, column=3, value=config.days_in_month(m)).font = NORMAL_FONT
        for col in range(1, 4):
            ws.cell(row=r, column=col).alignment = CENTER
            ws.cell(row=r, column=col).border = THIN_BORDER

    month_end = month_start + 13
    tbl2 = Table(displayName="tblMonthDays", ref=f"A{month_start + 1}:C{month_end}")
    tbl2.tableStyleInfo = TABLE_STYLE
    ws.add_table(tbl2)

    # Column widths
    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 22
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 14
    ws.column_dimensions["H"].width = 14

# ═══════════════════════════════════════════════════════════════════════════════
# DATA SHEETS (hidden tables)
# ═══════════════════════════════════════════════════════════════════════════════

def build_daily_data_sheet(wb: Workbook, config: WorkbookConfig):
    ws = wb.create_sheet("DailyData")
    ws.sheet_properties.tabColor = "808080"

    headers = [
        "EntryDate", "Month", "WardCode", "Admissions", "Discharges",
        "Deaths", "DeathsUnder24Hrs", "TransfersIn", "TransfersOut",
        "PrevRemaining", "Remaining", "EntryTimestamp"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)

    # Seed row (VBA calculates PrevRemaining and Remaining as VALUES)
    ws.cell(row=2, column=1, value="")  # EntryDate (VBA fills)
    ws.cell(row=2, column=2, value="")  # Month (VBA fills)
    ws.cell(row=2, column=3, value="")  # WardCode (VBA fills)
    ws.cell(row=2, column=4, value="")  # Admissions (VBA fills)
    ws.cell(row=2, column=5, value="")  # Discharges (VBA fills)
    ws.cell(row=2, column=6, value="")  # Deaths (VBA fills)
    ws.cell(row=2, column=7, value="")  # DeathsUnder24Hrs (VBA fills)
    ws.cell(row=2, column=8, value="")  # TransfersIn (VBA fills)
    ws.cell(row=2, column=9, value="")  # TransfersOut (VBA fills)
    ws.cell(row=2, column=10, value="")  # PrevRemaining (VBA calculates)
    ws.cell(row=2, column=11, value="")  # Remaining (VBA calculates)
    ws.cell(row=2, column=12, value="")  # EntryTimestamp (VBA fills)

    tbl = Table(displayName="tblDaily", ref="A1:L2")
    tbl.tableStyleInfo = TABLE_STYLE
    ws.add_table(tbl)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 12
    for c in "DEFGHIJ":
        ws.column_dimensions[c].width = 14
    ws.column_dimensions["K"].width = 12
    ws.column_dimensions["L"].width = 18


def build_admissions_sheet(wb: Workbook, config: WorkbookConfig):
    ws = wb.create_sheet("Admissions")
    ws.sheet_properties.tabColor = "808080"

    headers = [
        "AdmissionID", "AdmissionDate", "Month", "WardCode", "PatientID",
        "PatientName", "Age", "AgeUnit", "Sex", "NHIS", "EntryTimestamp"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    ws.cell(row=2, column=1, value="")

    tbl = Table(displayName="tblAdmissions", ref="A1:K2")
    tbl.tableStyleInfo = TABLE_STYLE
    ws.add_table(tbl)

    ws.column_dimensions["A"].width = 14
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 6
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 18


def build_deaths_data_sheet(wb: Workbook, config: WorkbookConfig):
    ws = wb.create_sheet("DeathsData")
    ws.sheet_properties.tabColor = "808080"

    headers = [
        "DeathID", "DateOfDeath", "Month", "WardCode", "FolderNumber",
        "NameOfDeceased", "Age", "AgeUnit", "Sex", "NHIS",
        "CauseOfDeath", "DeathWithin24Hrs", "EntryTimestamp"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    ws.cell(row=2, column=1, value="")

    tbl = Table(displayName="tblDeaths", ref="A1:M2")
    tbl.tableStyleInfo = TABLE_STYLE
    ws.add_table(tbl)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 8
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 10
    ws.column_dimensions["I"].width = 6
    ws.column_dimensions["J"].width = 14
    ws.column_dimensions["K"].width = 25
    ws.column_dimensions["L"].width = 16
    ws.column_dimensions["M"].width = 18


def build_transfers_sheet(wb: Workbook, config: WorkbookConfig):
    ws = wb.create_sheet("TransfersData")
    ws.sheet_properties.tabColor = "808080"

    headers = [
        "TransferID", "TransferDate", "Month", "FromWardCode",
        "ToWardCode", "PatientID", "PatientName", "EntryTimestamp"
    ]
    for col, h in enumerate(headers, 1):
        ws.cell(row=1, column=col, value=h)
    ws.cell(row=2, column=1, value="")

    tbl = Table(displayName="tblTransfers", ref="A1:H2")
    tbl.tableStyleInfo = TABLE_STYLE
    ws.add_table(tbl)


# ═══════════════════════════════════════════════════════════════════════════════
# WARD REPORT SHEETS (9 sheets, one per ward)
# ═══════════════════════════════════════════════════════════════════════════════

def build_ward_sheet(wb: Workbook, config: WorkbookConfig, ward):
    ws = wb.create_sheet(ward.name)

    # Tab colors: wards get distinct colors
    ward_colors = {
        "MW": "4472C4", "FW": "ED7D31", "CW": "A5A5A5",
        "BF": "FFC000", "BG": "5B9BD5", "BH": "70AD47",
        "NICU": "7030A0", "MAE": "FF0000", "FAE": "FF69B4",
    }
    ws.sheet_properties.tabColor = ward_colors.get(ward.code, "000000")

    current_row = 1
    for month_num in range(1, 13):
        month_name = config.MONTH_NAMES[month_num - 1]
        days = config.days_in_month(month_num)

        # ── Header block ─────────────────────────────────────────────
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=7)
        c = ws.cell(row=current_row, column=1, value="GHANA HEALTH SERVICE")
        c.font = HEADER_FONT
        c.alignment = CENTER
        current_row += 1

        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=7)
        c = ws.cell(row=current_row, column=1, value="DAILY BED UTILIZATION FORM")
        c.font = SUBHEADER_FONT
        c.alignment = CENTER
        current_row += 1

        # Hospital / Ward / Month row
        ws.cell(row=current_row, column=1, value="Hospital:").font = BOLD_FONT
        ws.cell(row=current_row, column=2, value=config.hospital_name).font = NORMAL_FONT
        ws.cell(row=current_row, column=4, value="Ward:").font = BOLD_FONT
        ws.cell(row=current_row, column=5, value=ward.name.upper()).font = NORMAL_FONT
        ws.cell(row=current_row, column=6, value="MONTH").font = BOLD_FONT
        ws.cell(row=current_row, column=7, value=month_name).font = NORMAL_FONT
        current_row += 1

        # Previous remaining
        ws.cell(row=current_row, column=1, value="Number of patients remaining as at last day of previous month").font = LABEL_FONT
        if month_num == 1:
            formula = (
                f'=IFERROR(INDEX(tblWardConfig[PrevYearRemaining],'
                f'MATCH("{ward.code}",tblWardConfig[WardCode],0)),0)'
            )
        else:
            prev_month = month_num - 1
            prev_last_day = config.days_in_month(prev_month)
            formula = (
                f'=IFERROR(SUMIFS(tblDaily[Remaining],'
                f'tblDaily[EntryDate],DATE({config.year},{prev_month},{prev_last_day}),'
                f'tblDaily[WardCode],"{ward.code}"),0)'
            )
        c = ws.cell(row=current_row, column=8, value=formula)
        c.font = BOLD_FONT
        c.fill = LIGHT_YELLOW_FILL
        c.border = THIN_BORDER
        prev_remaining_row = current_row
        current_row += 1

        # Bed complement
        ws.cell(row=current_row, column=1, value="Bed complement").font = LABEL_FONT
        bed_formula = (
            f'=IFERROR(INDEX(tblWardConfig[BedComplement],'
            f'MATCH("{ward.code}",tblWardConfig[WardCode],0)),0)'
        )
        c = ws.cell(row=current_row, column=8, value=bed_formula)
        c.font = BOLD_FONT
        c.fill = LIGHT_GREEN_FILL
        c.border = THIN_BORDER
        current_row += 1

        # ── Column headers ───────────────────────────────────────────
        col_headers = [
            "Day of\nthe\nMonth", "Admissions", "Discharges", "Deaths",
            "Deaths\n<24Hrs", "Transfers-\nIn", "Transfers-\nOut",
            "No. of Patients\nRemaining In\nWard"
        ]
        for col, header in enumerate(col_headers, 1):
            c = ws.cell(row=current_row, column=col, value=header)
            c.font = HEADER_FONT_WHITE
            c.fill = HEADER_FILL
            c.alignment = CENTER_WRAP
            c.border = THIN_BORDER
        header_row = current_row
        current_row += 1

        # ── Daily rows (1-31) ────────────────────────────────────────
        data_start = current_row
        for day in range(1, 32):
            ws.cell(row=current_row, column=1, value=day).font = NORMAL_FONT
            ws.cell(row=current_row, column=1).alignment = CENTER
            ws.cell(row=current_row, column=1).border = THIN_BORDER

            if day <= days:
                date_ref = f"DATE({config.year},{month_num},{day})"
                fields = ["Admissions", "Discharges", "Deaths",
                           "DeathsUnder24Hrs", "TransfersIn", "TransfersOut", "Remaining"]
                for col_idx, field_name in enumerate(fields, 2):
                    formula = (
                        f'=IFERROR(IF(SUMIFS(tblDaily[{field_name}],'
                        f'tblDaily[EntryDate],{date_ref},'
                        f'tblDaily[WardCode],"{ward.code}")=0,"",'
                        f'SUMIFS(tblDaily[{field_name}],'
                        f'tblDaily[EntryDate],{date_ref},'
                        f'tblDaily[WardCode],"{ward.code}")),"")'
                    )
                    c = ws.cell(row=current_row, column=col_idx, value=formula)
                    c.font = NORMAL_FONT
                    c.alignment = CENTER
                    c.border = THIN_BORDER
            else:
                for col_idx in range(2, 9):
                    c = ws.cell(row=current_row, column=col_idx)
                    c.fill = GRAY_FILL
                    c.border = THIN_BORDER

            current_row += 1

        # ── TOTAL row ────────────────────────────────────────────────
        c = ws.cell(row=current_row, column=1, value="TOTAL")
        c.font = BOLD_FONT
        c.alignment = CENTER
        c.fill = TOTAL_FILL
        c.border = THIN_BORDER

        total_fields = ["Admissions", "Discharges", "Deaths",
                        "DeathsUnder24Hrs", "TransfersIn", "TransfersOut"]
        for col_idx, field_name in enumerate(total_fields, 2):
            if field_name == "Admissions" and config.preferences.subtract_deaths_under_24hrs_from_admissions:
                # Adjusted admissions = Admissions - Deaths<24Hrs
                formula = (
                    f'=SUMIFS(tblDaily[Admissions],'
                    f'tblDaily[Month],{month_num},'
                    f'tblDaily[WardCode],"{ward.code}") - '
                    f'SUMIFS(tblDaily[DeathsUnder24Hrs],'
                    f'tblDaily[Month],{month_num},'
                    f'tblDaily[WardCode],"{ward.code}")'
                )
            else:
                formula = (
                    f'=SUMIFS(tblDaily[{field_name}],'
                    f'tblDaily[Month],{month_num},'
                    f'tblDaily[WardCode],"{ward.code}")'
                )
            c = ws.cell(row=current_row, column=col_idx, value=formula)
            c.font = BOLD_FONT
            c.alignment = CENTER
            c.fill = TOTAL_FILL
            c.border = THIN_BORDER

        # Patient Days (sum of daily remaining) - column 8
        pd_formula = (
            f'=SUMIFS(tblDaily[Remaining],'
            f'tblDaily[Month],{month_num},'
            f'tblDaily[WardCode],"{ward.code}")'
        )
        c = ws.cell(row=current_row, column=8, value=pd_formula)
        c.font = BOLD_FONT
        c.alignment = CENTER
        c.fill = TOTAL_FILL
        c.border = THIN_BORDER

        current_row += 2  # spacer

    # Column widths
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 10
    ws.column_dimensions["E"].width = 10
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 16

    # Print setup
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToPage = True


def build_emergency_combined_sheet(wb: Workbook, config: WorkbookConfig):
    """
    Build a combined Emergency sheet showing MAE and FAE data side-by-side.

    Layout: 12 monthly blocks with dual-column structure
    - Left columns (B-H): Male Emergency (MAE) data
    - Right columns (I-O): Female Emergency (FAE) data
    """
    # Find emergency wards
    mae_ward = next((w for w in config.WARDS if w.code == "MAE"), None)
    fae_ward = next((w for w in config.WARDS if w.code == "FAE"), None)

    if not mae_ward or not fae_ward:
        print("WARNING: MAE or FAE ward not found. Skipping Emergency combined sheet.")
        return

    # Create sheet
    ws = wb.create_sheet("Emergency")
    ws.sheet_properties.tabColor = "FF6600"  # Bright orange

    current_row = 1

    # Map field names to columns
    mae_fields = {
        2: "Admissions",       # Col B
        3: "Discharges",       # Col C
        4: "Deaths",           # Col D
        5: "DeathsUnder24Hrs", # Col E
        6: "TransfersIn",      # Col F
        7: "TransfersOut",     # Col G
        8: "Remaining"         # Col H
    }

    fae_fields = {
        9: "Admissions",       # Col I
        10: "Discharges",      # Col J
        11: "Deaths",          # Col K
        12: "DeathsUnder24Hrs",# Col L
        13: "TransfersIn",     # Col M
        14: "TransfersOut",    # Col N
        15: "Remaining"        # Col O
    }

    for month_num in range(1, 13):
        month_name = config.MONTH_NAMES[month_num - 1]

        # ── Header Block ──
        ws.merge_cells(f"A{current_row}:O{current_row}")
        c = ws.cell(row=current_row, column=1, value="GHANA HEALTH SERVICE")
        c.font = Font(name="Calibri", bold=True, size=14, color="1F4E79")
        c.alignment = CENTER
        current_row += 1

        ws.merge_cells(f"A{current_row}:O{current_row}")
        c = ws.cell(row=current_row, column=1, value="DAILY BED UTILIZATION FORM")
        c.font = Font(name="Calibri", bold=True, size=12)
        c.alignment = CENTER
        current_row += 1

        # ── Hospital/Ward/Month Info ──
        ws.cell(row=current_row, column=1, value="Hospital:").font = BOLD_FONT
        ws.cell(row=current_row, column=2, value=config.hospital_name).font = NORMAL_FONT

        ws.cell(row=current_row, column=5, value="Ward:").font = BOLD_FONT
        ws.cell(row=current_row, column=6, value="EMERGENCY (COMBINED)").font = BOLD_FONT

        ws.cell(row=current_row, column=13, value="MONTH").font = BOLD_FONT
        ws.cell(row=current_row, column=14, value=month_name).font = BOLD_FONT
        current_row += 1

        # ── Previous Remaining Row ──
        ws.cell(row=current_row, column=1, value="Number of patients remaining as at last day of previous month").font = BOLD_FONT

        # MAE previous remaining (col B)
        if month_num == 1:
            mae_prev_formula = f'=IFERROR(INDEX(tblWardConfig[PrevYearRemaining],MATCH("MAE",tblWardConfig[WardCode],0)),0)'
        else:
            prev_month = month_num - 1
            last_day = config.days_in_month(prev_month)
            mae_prev_formula = f'=IFERROR(SUMIFS(tblDaily[Remaining],tblDaily[EntryDate],DATE({config.year},{prev_month},{last_day}),tblDaily[WardCode],"MAE"),0)'
        ws.cell(row=current_row, column=2, value=mae_prev_formula).alignment = CENTER

        # FAE previous remaining (col I)
        if month_num == 1:
            fae_prev_formula = f'=IFERROR(INDEX(tblWardConfig[PrevYearRemaining],MATCH("FAE",tblWardConfig[WardCode],0)),0)'
        else:
            fae_prev_formula = f'=IFERROR(SUMIFS(tblDaily[Remaining],tblDaily[EntryDate],DATE({config.year},{prev_month},{last_day}),tblDaily[WardCode],"FAE"),0)'
        ws.cell(row=current_row, column=9, value=fae_prev_formula).alignment = CENTER

        current_row += 1

        # ── Bed Complement Row ──
        ws.cell(row=current_row, column=1, value="Bed complement").font = BOLD_FONT

        # MAE bed complement
        ws.cell(row=current_row, column=2, value=f'=IFERROR(INDEX(tblWardConfig[BedComplement],MATCH("MAE",tblWardConfig[WardCode],0)),0)').alignment = CENTER

        # FAE bed complement
        ws.cell(row=current_row, column=9, value=f'=IFERROR(INDEX(tblWardConfig[BedComplement],MATCH("FAE",tblWardConfig[WardCode],0)),0)').alignment = CENTER

        # Total emergency beds
        ws.cell(row=current_row, column=15, value=f'=B{current_row}+I{current_row}').alignment = CENTER
        ws.cell(row=current_row, column=15).font = BOLD_FONT

        current_row += 1

        # ── Column Headers (2 rows) ──
        # Row 1: Section headers
        ws.cell(row=current_row, column=1, value="Day").font = HEADER_FONT_WHITE
        ws.cell(row=current_row, column=1).fill = HEADER_FILL
        ws.cell(row=current_row, column=1).alignment = CENTER
        ws.cell(row=current_row, column=1).border = THIN_BORDER

        ws.merge_cells(f"B{current_row}:H{current_row}")
        c = ws.cell(row=current_row, column=2, value="MALE EMERGENCY")
        c.font = HEADER_FONT_WHITE
        c.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")  # Blue
        c.alignment = CENTER
        c.border = THIN_BORDER

        ws.merge_cells(f"I{current_row}:O{current_row}")
        c = ws.cell(row=current_row, column=9, value="FEMALE EMERGENCY")
        c.font = HEADER_FONT_WHITE
        c.fill = PatternFill(start_color="ED7D31", end_color="ED7D31", fill_type="solid")  # Orange
        c.alignment = CENTER
        c.border = THIN_BORDER

        # Total Header
        c = ws.cell(row=current_row, column=16, value="TOTAL")
        c.font = HEADER_FONT_WHITE
        c.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")  # Green
        c.alignment = CENTER
        c.border = THIN_BORDER

        current_row += 1

        # Row 2: Field headers
        field_headers = [
            (1, "Day"),
            # Male Emergency columns
            (2, "Adm"), (3, "Dis"), (4, "Dth"), (5, "D<24"),
            (6, "TrIn"), (7, "TrOut"), (8, "Rem"),
            # Female Emergency columns
            (9, "Adm"), (10, "Dis"), (11, "Dth"), (12, "D<24"),
            (13, "TrIn"), (14, "TrOut"), (15, "Rem"),
            # Total Column
            (16, "Tot\nRem")
        ]

        for col_num, header_text in field_headers:
            c = ws.cell(row=current_row, column=col_num, value=header_text)
            c.font = HEADER_FONT_WHITE
            c.fill = HEADER_FILL
            c.alignment = CENTER
            c.border = THIN_BORDER

        current_row += 1

        # ── Daily Rows (1-31) ──
        days_in_month = config.days_in_month(month_num)

        for day in range(1, 32):
            # Day number (col A)
            c = ws.cell(row=current_row, column=1, value=day)
            c.alignment = CENTER
            c.border = THIN_BORDER

            if day <= days_in_month:
                # MAE data (cols B-H)
                for col_num, field_name in mae_fields.items():
                    formula = (
                        f'=IFERROR(IF(SUMIFS(tblDaily[{field_name}],'
                        f'tblDaily[EntryDate],DATE({config.year},{month_num},{day}),'
                        f'tblDaily[WardCode],"MAE")=0,"",'
                        f'SUMIFS(tblDaily[{field_name}],'
                        f'tblDaily[EntryDate],DATE({config.year},{month_num},{day}),'
                        f'tblDaily[WardCode],"MAE")),"")'
                    )
                    c = ws.cell(row=current_row, column=col_num, value=formula)
                    c.alignment = CENTER
                    c.border = THIN_BORDER

                # FAE data (cols I-O)
                for col_num, field_name in fae_fields.items():
                    formula = (
                        f'=IFERROR(IF(SUMIFS(tblDaily[{field_name}],'
                        f'tblDaily[EntryDate],DATE({config.year},{month_num},{day}),'
                        f'tblDaily[WardCode],"FAE")=0,"",'
                        f'SUMIFS(tblDaily[{field_name}],'
                        f'tblDaily[EntryDate],DATE({config.year},{month_num},{day}),'
                        f'tblDaily[WardCode],"FAE")),"")'
                    )
                    c = ws.cell(row=current_row, column=col_num, value=formula)
                    c.alignment = CENTER
                    c.border = THIN_BORDER

                # Total Remaining (Col P = Col H + Col O)
                # Check directly if cells have numbers to avoid summing text
                mae_rem = f"H{current_row}"
                fae_rem = f"O{current_row}"
                formula = f'=IF(AND(ISNUMBER({mae_rem}), ISNUMBER({fae_rem})), {mae_rem} + {fae_rem}, IF(ISNUMBER({mae_rem}), {mae_rem}, IF(ISNUMBER({fae_rem}), {fae_rem}, "")))'
                
                # Simpler: just sum them, treating text as 0 is standard SUM behavior but + operator needs numbers
                # Actually, our formulas return "" if 0. 
                # Better formula: =N(H_row) + N(O_row) but we want "" if both are empty.
                # Let's use: =IF(AND(H="" , O=""), "", N(H)+N(O))
                formula = f'=IF(AND({mae_rem}="", {fae_rem}=""), "", N({mae_rem}) + N({fae_rem}))'

                c = ws.cell(row=current_row, column=16, value=formula)
                c.alignment = CENTER
                c.border = THIN_BORDER
                c.font = BOLD_FONT

            else:
                # Gray out invalid days
                for col_num in range(2, 17):  # Cols B-P
                    c = ws.cell(row=current_row, column=col_num)
                    c.fill = GRAY_FILL
                    c.border = THIN_BORDER

            current_row += 1

        # ── TOTAL Row ──
        c = ws.cell(row=current_row, column=1, value="TOTAL")
        c.font = BOLD_FONT
        c.alignment = CENTER
        c.fill = TOTAL_FILL
        c.border = THIN_BORDER

        # MAE totals (cols B-H)
        for col_num, field_name in mae_fields.items():
            if field_name == "Admissions" and config.preferences.subtract_deaths_under_24hrs_from_admissions:
                formula = (
                    f'=SUMIFS(tblDaily[Admissions],tblDaily[Month],{month_num},tblDaily[WardCode],"MAE") - '
                    f'SUMIFS(tblDaily[DeathsUnder24Hrs],tblDaily[Month],{month_num},tblDaily[WardCode],"MAE")'
                )
            else:
                formula = f'=SUMIFS(tblDaily[{field_name}],tblDaily[Month],{month_num},tblDaily[WardCode],"MAE")'

            c = ws.cell(row=current_row, column=col_num, value=formula)
            c.font = BOLD_FONT
            c.alignment = CENTER
            c.fill = TOTAL_FILL
            c.border = THIN_BORDER

        # FAE totals (cols I-O)
        for col_num, field_name in fae_fields.items():
            if field_name == "Admissions" and config.preferences.subtract_deaths_under_24hrs_from_admissions:
                formula = (
                    f'=SUMIFS(tblDaily[Admissions],tblDaily[Month],{month_num},tblDaily[WardCode],"FAE") - '
                    f'SUMIFS(tblDaily[DeathsUnder24Hrs],tblDaily[Month],{month_num},tblDaily[WardCode],"FAE")'
                )
            else:
                formula = f'=SUMIFS(tblDaily[{field_name}],tblDaily[Month],{month_num},tblDaily[WardCode],"FAE")'

            c = ws.cell(row=current_row, column=col_num, value=formula)
            c.font = BOLD_FONT
            c.alignment = CENTER
            c.fill = TOTAL_FILL
            c.border = THIN_BORDER

        # Total Remaining Sum (Col P)
        # It's sum of daily remainings (Patient Days)
        # = SUM(Col P daily rows) OR Sum of MAE Rem Total + FAE Rem Total
        mae_rem_tot = f"H{current_row}"
        fae_rem_tot = f"O{current_row}"
        formula = f'={mae_rem_tot} + {fae_rem_tot}'
        c = ws.cell(row=current_row, column=16, value=formula)
        c.font = BOLD_FONT
        c.alignment = CENTER
        c.fill = TOTAL_FILL
        c.border = THIN_BORDER

        current_row += 2  # Spacer before next month

    # Column widths
    ws.column_dimensions["A"].width = 5   # Day
    ws.column_dimensions["B"].width = 6   # MAE Adm
    ws.column_dimensions["C"].width = 6   # MAE Dis
    ws.column_dimensions["D"].width = 6   # MAE Dth
    ws.column_dimensions["E"].width = 6   # MAE D<24
    ws.column_dimensions["F"].width = 6   # MAE TrIn
    ws.column_dimensions["G"].width = 6   # MAE TrOut
    ws.column_dimensions["H"].width = 6   # MAE Rem
    ws.column_dimensions["I"].width = 6   # FAE Adm
    ws.column_dimensions["J"].width = 6   # FAE Dis
    ws.column_dimensions["K"].width = 6   # FAE Dth
    ws.column_dimensions["L"].width = 6   # FAE D<24
    ws.column_dimensions["M"].width = 6   # FAE TrIn
    ws.column_dimensions["N"].width = 6   # FAE TrOut
    ws.column_dimensions["O"].width = 6   # FAE Rem
    ws.column_dimensions["P"].width = 8   # Tot Rem

    # Print setup
    ws.page_setup.orientation = "landscape"  # Wider sheet
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToPage = True


# ═══════════════════════════════════════════════════════════════════════════════
# MONTHLY SUMMARY SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def build_monthly_summary_sheet(wb: Workbook, config: WorkbookConfig):
    ws = wb.create_sheet("Monthly Summary")
    ws.sheet_properties.tabColor = "1F4E79"

    current_row = 1
    for month_num in range(1, 13):
        month_name = config.MONTH_NAMES[month_num - 1]
        days = config.days_in_month(month_num)

        # Header
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=16)
        c = ws.cell(row=current_row, column=1, value="GHANA HEALTH SERVICE")
        c.font = HEADER_FONT
        c.alignment = CENTER
        current_row += 1

        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=16)
        c = ws.cell(row=current_row, column=1,
                    value=f"MONTHLY BED UTILIZATION FORM - {month_name}, {config.year}")
        c.font = SUBHEADER_FONT
        c.alignment = CENTER
        current_row += 1

        # Column headers
        col_headers = [
            "WARD", "Patients on bed\nat the beginning\nof the month",
            "Bed Complement", "Admissions", "Discharges", "Deaths",
            "Deaths\n<24Hrs", "Patient\nDays", "Transfer In", "Transfer\nOut",
            "Average\nDaily Bed\nOccupancy", "Average\nLength\nof Stay",
            "Bed\nTurnover\nInterval", "Bed\nTurnover\nRate",
            "Percentage\nof\nOccupancy", "Death\nRate"
        ]
        for col, header in enumerate(col_headers, 1):
            c = ws.cell(row=current_row, column=col, value=header)
            c.font = HEADER_FONT_WHITE
            c.fill = HEADER_FILL
            c.alignment = CENTER_WRAP
            c.border = THIN_BORDER
        header_row = current_row
        current_row += 1

        # ── Ward rows ────────────────────────────────────────────────
        for ward in config.WARDS:
            r = current_row
            wc = ward.code

            # Col A: Ward name
            ws.cell(row=r, column=1, value=ward.name).font = BOLD_FONT
            ws.cell(row=r, column=1).border = THIN_BORDER

            # Col B: Patients at beginning of month
            if month_num == 1:
                f_beg = (
                    f'=IFERROR(INDEX(tblWardConfig[PrevYearRemaining],'
                    f'MATCH("{wc}",tblWardConfig[WardCode],0)),0)'
                )
            else:
                pm = month_num - 1
                pld = config.days_in_month(pm)
                f_beg = (
                    f'=IFERROR(SUMIFS(tblDaily[Remaining],'
                    f'tblDaily[EntryDate],DATE({config.year},{pm},{pld}),'
                    f'tblDaily[WardCode],"{wc}"),0)'
                )
            ws.cell(row=r, column=2, value=f_beg).font = NORMAL_FONT

            # Col C: Bed Complement
            f_bc = f'=IFERROR(INDEX(tblWardConfig[BedComplement],MATCH("{wc}",tblWardConfig[WardCode],0)),0)'
            ws.cell(row=r, column=3, value=f_bc).font = NORMAL_FONT

            # Col D-F,G: Admissions, Discharges, Deaths, Deaths<24Hrs
            daily_fields = {
                4: "Admissions", 5: "Discharges", 6: "Deaths", 7: "DeathsUnder24Hrs"
            }
            for col_num, field in daily_fields.items():
                if field == "Admissions" and config.preferences.subtract_deaths_under_24hrs_from_admissions:
                    # Adjusted admissions = Admissions - Deaths<24Hrs
                    f = (f'=SUMIFS(tblDaily[Admissions],tblDaily[Month],{month_num},tblDaily[WardCode],"{wc}") - '
                         f'SUMIFS(tblDaily[DeathsUnder24Hrs],tblDaily[Month],{month_num},tblDaily[WardCode],"{wc}")')
                else:
                    f = f'=SUMIFS(tblDaily[{field}],tblDaily[Month],{month_num},tblDaily[WardCode],"{wc}")'
                ws.cell(row=r, column=col_num, value=f).font = NORMAL_FONT

            # Col H: Patient Days
            f_pd = f'=SUMIFS(tblDaily[Remaining],tblDaily[Month],{month_num},tblDaily[WardCode],"{wc}")'
            ws.cell(row=r, column=8, value=f_pd).font = NORMAL_FONT

            # Col I-J: Transfers In, Transfers Out
            f_ti = f'=SUMIFS(tblDaily[TransfersIn],tblDaily[Month],{month_num},tblDaily[WardCode],"{wc}")'
            f_to = f'=SUMIFS(tblDaily[TransfersOut],tblDaily[Month],{month_num},tblDaily[WardCode],"{wc}")'
            ws.cell(row=r, column=9, value=f_ti).font = NORMAL_FONT
            ws.cell(row=r, column=10, value=f_to).font = NORMAL_FONT

            # Col K: Average Daily Bed Occupancy = Patient Days / Days
            bc = get_column_letter(8)
            ws.cell(row=r, column=11, value=f'=IFERROR({bc}{r}/{days},0)').font = NORMAL_FONT

            # Col L: Average Length of Stay = Patient Days / (Discharges + Deaths)
            ws.cell(row=r, column=12, value=f'=IFERROR({bc}{r}/(E{r}+F{r}),0)').font = NORMAL_FONT

            # Col M: Bed Turnover Interval = (BC*Days - PD) / (Disch + Deaths)
            ws.cell(row=r, column=13, value=f'=IFERROR((C{r}*{days}-H{r})/(E{r}+F{r}),0)').font = NORMAL_FONT

            # Col N: Bed Turnover Rate = (Disch + Deaths) / BC
            ws.cell(row=r, column=14, value=f'=IFERROR((E{r}+F{r})/C{r},0)').font = NORMAL_FONT

            # Col O: % Occupancy = (PD / (BC * Days)) * 100
            ws.cell(row=r, column=15, value=f'=IFERROR((H{r}/(C{r}*{days}))*100,0)').font = NORMAL_FONT

            # Col P: Death Rate = Deaths / (Admissions + Patients at beginning) * 100
            ws.cell(row=r, column=16, value=f'=IFERROR((F{r}/(D{r}+B{r}))*100,0)').font = NORMAL_FONT

            # Format numbers
            for col in range(11, 17):
                ws.cell(row=r, column=col).number_format = '0.00'

            # Borders
            for col in range(1, 17):
                ws.cell(row=r, column=col).alignment = CENTER
                ws.cell(row=r, column=col).border = THIN_BORDER

            current_row += 1

        # ── EMERGENCY TOTAL REMAINING row ────────────────────────────
        if config.preferences.show_emergency_total_remaining:
            r = current_row
            ws.cell(row=r, column=1, value="EMERGENCY TOTAL REMAINING").font = BOLD_FONT
            ws.cell(row=r, column=1).fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

            # Calculate last day of month
            last_day = config.days_in_month(month_num)
            date_formula = f'DATE({config.year},{month_num},{last_day})'

            # Sum remaining for all emergency wards on last day
            emer_wards = [w for w in config.WARDS if w.is_emergency]
            parts = []
            for ew in emer_wards:
                parts.append(f'SUMIFS(tblDaily[Remaining],tblDaily[EntryDate],{date_formula},tblDaily[WardCode],"{ew.code}")')

            # Column B: Total remaining at end of month
            ws.cell(row=r, column=2, value=f'={"+".join(parts)}').font = BOLD_FONT

            # Columns C-P: Empty
            for col in range(3, 17):
                ws.cell(row=r, column=col, value="")
                ws.cell(row=r, column=col).fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")

            # Borders
            for col in range(1, 17):
                ws.cell(row=r, column=col).alignment = CENTER
                ws.cell(row=r, column=col).border = THIN_BORDER

            current_row += 1

        # ── TOTAL row ────────────────────────────────────────────────
        r = current_row
        ws.cell(row=r, column=1, value="TOTAL").font = BOLD_FONT
        ws.cell(row=r, column=1).fill = TOTAL_FILL

        # Sum columns B through J across the ward rows
        emergency_offset = 1 if config.preferences.show_emergency_total_remaining else 0
        first_ward_row = r - len(config.WARDS) - emergency_offset
        last_ward_row = r - 1 - emergency_offset
        for col in range(2, 11):
            cl = get_column_letter(col)
            ws.cell(row=r, column=col,
                    value=f'=SUM({cl}{first_ward_row}:{cl}{last_ward_row})').font = BOLD_FONT

        # KPI totals use total values
        ws.cell(row=r, column=11, value=f'=IFERROR(H{r}/{days},0)').font = BOLD_FONT
        ws.cell(row=r, column=12, value=f'=IFERROR(H{r}/(E{r}+F{r}),0)').font = BOLD_FONT
        ws.cell(row=r, column=13, value=f'=IFERROR((C{r}*{days}-H{r})/(E{r}+F{r}),0)').font = BOLD_FONT
        ws.cell(row=r, column=14, value=f'=IFERROR((E{r}+F{r})/C{r},0)').font = BOLD_FONT
        ws.cell(row=r, column=15, value=f'=IFERROR((H{r}/(C{r}*{days}))*100,0)').font = BOLD_FONT
        ws.cell(row=r, column=16, value=f'=IFERROR((F{r}/(D{r}+B{r}))*100,0)').font = BOLD_FONT

        for col in range(1, 17):
            ws.cell(row=r, column=col).alignment = CENTER
            ws.cell(row=r, column=col).border = THIN_BORDER
            ws.cell(row=r, column=col).fill = TOTAL_FILL
        for col in range(11, 17):
            ws.cell(row=r, column=col).number_format = '0.00'
        current_row += 1

        # ── Emergency subtotal row ───────────────────────────────────
        r = current_row
        ws.cell(row=r, column=1, value="Emergency").font = BOLD_FONT
        ws.cell(row=r, column=1).fill = LIGHT_YELLOW_FILL

        emer_wards = [w for w in config.WARDS if w.is_emergency]
        for col in range(2, 11):
            parts = []
            for ew in emer_wards:
                # Find the row for this emergency ward
                ew_idx = config.WARDS.index(ew)
                ew_row = first_ward_row + ew_idx
                parts.append(f'{get_column_letter(col)}{ew_row}')
            ws.cell(row=r, column=col, value=f'={"+".join(parts)}').font = BOLD_FONT

        ws.cell(row=r, column=11, value=f'=IFERROR(H{r}/{days},0)').font = BOLD_FONT
        ws.cell(row=r, column=12, value=f'=IFERROR(H{r}/(E{r}+F{r}),0)').font = BOLD_FONT
        ws.cell(row=r, column=13, value=f'=IFERROR((C{r}*{days}-H{r})/(E{r}+F{r}),0)').font = BOLD_FONT
        ws.cell(row=r, column=14, value=f'=IFERROR((E{r}+F{r})/C{r},0)').font = BOLD_FONT
        ws.cell(row=r, column=15, value=f'=IFERROR((H{r}/(C{r}*{days}))*100,0)').font = BOLD_FONT
        ws.cell(row=r, column=16, value=f'=IFERROR((F{r}/(D{r}+B{r}))*100,0)').font = BOLD_FONT

        for col in range(1, 17):
            ws.cell(row=r, column=col).alignment = CENTER
            ws.cell(row=r, column=col).border = THIN_BORDER
            ws.cell(row=r, column=col).fill = LIGHT_YELLOW_FILL
        for col in range(11, 17):
            ws.cell(row=r, column=col).number_format = '0.00'

        current_row += 2  # spacer before next month

    # Column widths
    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 14
    for i in range(4, 11):
        ws.column_dimensions[get_column_letter(i)].width = 12
    for i in range(11, 17):
        ws.column_dimensions[get_column_letter(i)].width = 12

    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToPage = True


# ═══════════════════════════════════════════════════════════════════════════════
# AGES SUMMARY SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def build_ages_summary_sheet(wb: Workbook, config: WorkbookConfig):
    ws = wb.create_sheet("Ages Summary")
    ws.sheet_properties.tabColor = "7030A0"

    # Layout: 12 horizontal sections (one per month), each occupying ~7 columns
    # For each month section:
    # Row 1: Month name
    # Row 2: Category headers (both ins and non ins | NON INS)
    # Row 3: Male | Female | Male | Female subheaders
    # Row 4-15: Age group data
    # Row 16: Total

    SECTION_WIDTH = 7  # columns per month section (age_label + 6 data cols)

    # Row 1: Month headers
    for m in range(1, 13):
        start_col = 1 + (m - 1) * SECTION_WIDTH
        ws.merge_cells(start_row=1, start_column=start_col,
                       end_row=1, end_column=start_col + SECTION_WIDTH - 1)
        c = ws.cell(row=1, column=start_col, value=config.MONTH_NAMES[m - 1])
        c.font = SUBHEADER_FONT
        c.alignment = CENTER
        c.fill = HEADER_FILL
        c.font = HEADER_FONT_WHITE

    # Row 2: Category headers
    for m in range(1, 13):
        sc = 1 + (m - 1) * SECTION_WIDTH
        # Col 0: blank (age group label)
        # Cols 1-2: "both ins and non ins" Male/Female
        ws.merge_cells(start_row=2, start_column=sc + 1, end_row=2, end_column=sc + 2)
        c = ws.cell(row=2, column=sc + 1, value="both ins and non ins")
        c.font = LABEL_FONT
        c.alignment = CENTER
        c.fill = LIGHT_BLUE_FILL

        # Cols 3-4: "NON INS" Male/Female
        ws.merge_cells(start_row=2, start_column=sc + 3, end_row=2, end_column=sc + 4)
        c = ws.cell(row=2, column=sc + 3, value="NON INS")
        c.font = LABEL_FONT
        c.alignment = CENTER
        c.fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")

        # Cols 5-6: "INS" (derived = total - non-ins, but we use COUNTIFS)
        # Actually from the screenshot, INS is shown separately
        # Let me add INS columns too
        ws.merge_cells(start_row=2, start_column=sc + 5, end_row=2, end_column=sc + 6)
        c = ws.cell(row=2, column=sc + 5, value="INS")
        c.font = LABEL_FONT
        c.alignment = CENTER
        c.fill = LIGHT_GREEN_FILL

    # Row 3: Male/Female subheaders
    for m in range(1, 13):
        sc = 1 + (m - 1) * SECTION_WIDTH
        ws.cell(row=3, column=sc, value="Age Group").font = BOLD_FONT
        ws.cell(row=3, column=sc).alignment = CENTER
        ws.cell(row=3, column=sc).border = THIN_BORDER
        for offset, label in [(1, "Male"), (2, "FEMALE"),
                               (3, "MALE"), (4, "FRMALE"),
                               (5, "MALE"), (6, "FRMALE")]:
            c = ws.cell(row=3, column=sc + offset, value=label)
            c.font = BOLD_FONT
            c.alignment = CENTER
            c.border = THIN_BORDER

    # Rows 4-15: Age group data with COUNTIFS formulas
    for m in range(1, 13):
        sc = 1 + (m - 1) * SECTION_WIDTH

        for ag_idx, (ag_label, age_unit, age_min, age_max) in enumerate(config.AGE_GROUPS):
            r = 4 + ag_idx
            ws.cell(row=r, column=sc, value=ag_label).font = NORMAL_FONT
            ws.cell(row=r, column=sc).alignment = CENTER
            ws.cell(row=r, column=sc).border = THIN_BORDER

            # Build COUNTIFS for each combination
            # Both ins and non-ins Male (col sc+1)
            base = (
                f'COUNTIFS(tblAdmissions[Month],{m},'
                f'tblAdmissions[AgeUnit],"{age_unit}",'
                f'tblAdmissions[Age],">="&{age_min},'
                f'tblAdmissions[Age],"<="&{age_max},'
            )
            # Total Male
            ws.cell(row=r, column=sc + 1,
                    value=f'={base}tblAdmissions[Sex],"M")').font = NORMAL_FONT
            # Total Female
            ws.cell(row=r, column=sc + 2,
                    value=f'={base}tblAdmissions[Sex],"F")').font = NORMAL_FONT
            # Non-Insured Male
            ws.cell(row=r, column=sc + 3,
                    value=f'={base}tblAdmissions[Sex],"M",tblAdmissions[NHIS],"Non-Insured")').font = NORMAL_FONT
            # Non-Insured Female
            ws.cell(row=r, column=sc + 4,
                    value=f'={base}tblAdmissions[Sex],"F",tblAdmissions[NHIS],"Non-Insured")').font = NORMAL_FONT
            # Insured Male
            ws.cell(row=r, column=sc + 5,
                    value=f'={base}tblAdmissions[Sex],"M",tblAdmissions[NHIS],"Insured")').font = NORMAL_FONT
            # Insured Female
            ws.cell(row=r, column=sc + 6,
                    value=f'={base}tblAdmissions[Sex],"F",tblAdmissions[NHIS],"Insured")').font = NORMAL_FONT

            for col_off in range(1, 7):
                ws.cell(row=r, column=sc + col_off).alignment = CENTER
                ws.cell(row=r, column=sc + col_off).border = THIN_BORDER

        # Total row
        total_row = 4 + len(config.AGE_GROUPS)
        ws.cell(row=total_row, column=sc, value="Total").font = BOLD_FONT
        ws.cell(row=total_row, column=sc).alignment = CENTER
        ws.cell(row=total_row, column=sc).fill = TOTAL_FILL
        ws.cell(row=total_row, column=sc).border = THIN_BORDER
        for col_off in range(1, 7):
            cl = get_column_letter(sc + col_off)
            ws.cell(row=total_row, column=sc + col_off,
                    value=f'=SUM({cl}4:{cl}{total_row - 1})').font = BOLD_FONT
            ws.cell(row=total_row, column=sc + col_off).alignment = CENTER
            ws.cell(row=total_row, column=sc + col_off).fill = TOTAL_FILL
            ws.cell(row=total_row, column=sc + col_off).border = THIN_BORDER

    # Set column widths
    for m in range(1, 13):
        sc = 1 + (m - 1) * SECTION_WIDTH
        ws.column_dimensions[get_column_letter(sc)].width = 8
        for off in range(1, 7):
            ws.column_dimensions[get_column_letter(sc + off)].width = 10


# ═══════════════════════════════════════════════════════════════════════════════
# DEATHS REPORT SHEET (VBA-refreshed, but we set up the structure)
# ═══════════════════════════════════════════════════════════════════════════════

def build_deaths_report_sheet(wb: Workbook, config: WorkbookConfig):
    ws = wb.create_sheet("Deaths Report")
    ws.sheet_properties.tabColor = "C00000"

    # Create 12 month sections with headers
    current_row = 1
    for month_num in range(1, 13):
        month_name = config.MONTH_NAMES[month_num - 1]

        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=8)
        c = ws.cell(row=current_row, column=1,
                    value=f"DEATHS FOR {month_name}")
        c.font = SUBHEADER_FONT
        c.alignment = CENTER
        current_row += 1

        headers = ["S/N", "Folder Number", "Date of Death",
                   "Name of Deceased", "Age", "Sex", "Ward", "NHIS"]
        for col, h in enumerate(headers, 1):
            c = ws.cell(row=current_row, column=col, value=h)
            c.font = HEADER_FONT_WHITE
            c.fill = HEADER_FILL
            c.alignment = CENTER
            c.border = THIN_BORDER
        current_row += 1

        # Leave space for data (VBA will populate this)
        # Reserve 40 rows per month
        current_row += 40

    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 8
    ws.column_dimensions["F"].width = 6
    ws.column_dimensions["G"].width = 10
    ws.column_dimensions["H"].width = 10


# ═══════════════════════════════════════════════════════════════════════════════
# COD SUMMARY SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def build_cod_summary_sheet(wb: Workbook, config: WorkbookConfig):
    ws = wb.create_sheet("COD Summary")
    ws.sheet_properties.tabColor = "C00000"

    ws.merge_cells("A1:N1")
    c = ws.cell(row=1, column=1, value="CAUSE OF DEATH SUMMARY")
    c.font = HEADER_FONT
    c.alignment = CENTER

    # Headers: Cause of Death | Jan | Feb | ... | Dec | Total
    headers = ["Cause of Death"] + list(config.MONTH_NAMES) + ["TOTAL"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = HEADER_FONT_WHITE
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = THIN_BORDER

    # VBA will populate rows 3+ with distinct causes and COUNTIFS
    # Leave as placeholder
    ws.cell(row=3, column=1, value="(Click 'Refresh Reports' to populate)").font = NORMAL_FONT

    ws.column_dimensions["A"].width = 30
    for col in range(2, 15):
        ws.column_dimensions[get_column_letter(col)].width = 10


# ═══════════════════════════════════════════════════════════════════════════════
# STATEMENT OF INPATIENT SHEET (Yearly Summary)
# ═══════════════════════════════════════════════════════════════════════════════

def build_statement_of_inpatient_sheet(wb: Workbook, config: WorkbookConfig):
    ws = wb.create_sheet("Statement of Inpatient")
    ws.sheet_properties.tabColor = "0000FF"

    ws.merge_cells("A1:P1")
    c = ws.cell(row=1, column=1, value="GHANA HEALTH SERVICE - STATEMENT OF INPATIENT (YEARLY SUMMARY)")
    c.font = HEADER_FONT
    c.alignment = CENTER

    # Headers same as Monthly Summary
    col_headers = [
        "WARD", "Patients on bed\nat start of year",
        "Bed Complement", "Admissions", "Discharges", "Deaths",
        "Deaths\n<24Hrs", "Patient\nDays", "Transfer In", "Transfer\nOut",
        "Average\nDaily Bed\nOccupancy", "Average\nLength\nof Stay",
        "Bed\nTurnover\nInterval", "Bed\nTurnover\nRate",
        "Percentage\nof\nOccupancy", "Death\nRate"
    ]
    for col, header in enumerate(col_headers, 1):
        c = ws.cell(row=3, column=col, value=header)
        c.font = HEADER_FONT_WHITE
        c.fill = HEADER_FILL
        c.alignment = CENTER_WRAP
        c.border = THIN_BORDER
    
    current_row = 4
    days_in_year = 365 + (1 if config.year % 4 == 0 else 0)

    for ward in config.WARDS:
        r = current_row
        wc = ward.code
        
        # A: Name
        ws.cell(row=r, column=1, value=ward.name).font = BOLD_FONT
        
        # B: Start of year
        f_beg = (
            f'=IFERROR(INDEX(tblWardConfig[PrevYearRemaining],'
            f'MATCH("{wc}",tblWardConfig[WardCode],0)),0)'
        )
        ws.cell(row=r, column=2, value=f_beg).font = NORMAL_FONT
        
        # C: BC
        f_bc = f'=IFERROR(INDEX(tblWardConfig[BedComplement],MATCH("{wc}",tblWardConfig[WardCode],0)),0)'
        ws.cell(row=r, column=3, value=f_bc).font = NORMAL_FONT
        
        # D-G, I-J: Sums
        if config.preferences.subtract_deaths_under_24hrs_from_admissions:
            adm_formula = f'=SUMIFS(tblDaily[Admissions],tblDaily[WardCode],"{wc}") - SUMIFS(tblDaily[DeathsUnder24Hrs],tblDaily[WardCode],"{wc}")'
        else:
            adm_formula = f'=SUMIFS(tblDaily[Admissions],tblDaily[WardCode],"{wc}")'

        ws.cell(row=r, column=4, value=adm_formula)
        ws.cell(row=r, column=5, value=f'=SUMIFS(tblDaily[Discharges],tblDaily[WardCode],"{wc}")')
        ws.cell(row=r, column=6, value=f'=SUMIFS(tblDaily[Deaths],tblDaily[WardCode],"{wc}")')
        ws.cell(row=r, column=7, value=f'=SUMIFS(tblDaily[DeathsUnder24Hrs],tblDaily[WardCode],"{wc}")')
        ws.cell(row=r, column=8, value=f'=SUMIFS(tblDaily[Remaining],tblDaily[WardCode],"{wc}")')
        ws.cell(row=r, column=9, value=f'=SUMIFS(tblDaily[TransfersIn],tblDaily[WardCode],"{wc}")')
        ws.cell(row=r, column=10, value=f'=SUMIFS(tblDaily[TransfersOut],tblDaily[WardCode],"{wc}")')
        
        # KPIs
        bc_Ref = f"C{r}"
        pd_Ref = f"H{r}"
        adm_Ref = f"D{r}"
        dis_Ref = f"E{r}"
        dth_Ref = f"F{r}"
        beg_Ref = f"B{r}"
        
        # Av Occ = PD / 365
        ws.cell(row=r, column=11, value=f'=IFERROR({pd_Ref}/{days_in_year},0)')
        # Av LOS = PD / (Dis + Dth)
        ws.cell(row=r, column=12, value=f'=IFERROR({pd_Ref}/({dis_Ref}+{dth_Ref}),0)')
        # Turn Int = (BC*365 - PD) / (Dis + Dth)
        ws.cell(row=r, column=13, value=f'=IFERROR(({bc_Ref}*{days_in_year}-{pd_Ref})/({dis_Ref}+{dth_Ref}),0)')
        # Turn Rate = (Dis + Dth) / BC
        ws.cell(row=r, column=14, value=f'=IFERROR(({dis_Ref}+{dth_Ref})/{bc_Ref},0)')
        # % Occ = (PD / (BC*365)) * 100
        ws.cell(row=r, column=15, value=f'=IFERROR(({pd_Ref}/({bc_Ref}*{days_in_year}))*100,0)')
        # Death Rate
        ws.cell(row=r, column=16, value=f'=IFERROR(({dth_Ref}/({adm_Ref}+{beg_Ref}))*100,0)')
        
        current_row += 1

    # Format
    for r in range(4, current_row):
        for c in range(1, 17):
            ws.cell(row=r, column=c).border = THIN_BORDER
        for c in range(11, 17):
            ws.cell(row=r, column=c).number_format = '0.00'
            
    ws.column_dimensions["A"].width = 18
    for i in range(2, 17):
        ws.column_dimensions[get_column_letter(i)].width = 12


# ═══════════════════════════════════════════════════════════════════════════════
# NON-INSURED REPORT SHEET
# ═══════════════════════════════════════════════════════════════════════════════

def build_non_insured_report_sheet(wb: Workbook, config: WorkbookConfig):
    ws = wb.create_sheet("Non-Insured Report")
    ws.sheet_properties.tabColor = "FFA500"

    ws.merge_cells("A1:J1")
    c = ws.cell(row=1, column=1, value="NON-INSURED ATTENDANTS REPORT")
    c.font = HEADER_FONT
    c.alignment = CENTER

    headers = ["S/N", "Date", "Month", "Ward", "Patient ID", "Name", "Age", "Sex", "Amount", "Status"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=2, column=col, value=h)
        c.font = HEADER_FONT_WHITE
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = THIN_BORDER

    ws.cell(row=3, column=1, value="(Click 'Refresh Reports' to populate)").font = NORMAL_FONT
    
    ws.column_dimensions["A"].width = 6
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 10
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 15
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 6
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 12


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN BUILD FUNCTION
# ═══════════════════════════════════════════════════════════════════════════════

def build_structure(config: WorkbookConfig, output_path: str):
    wb = Workbook()

    # 1. Control sheet (landing page)
    build_control_sheet(wb, config)

    # 2. Data sheets (hidden tables)
    build_daily_data_sheet(wb, config)
    build_admissions_sheet(wb, config)
    build_deaths_data_sheet(wb, config)
    build_transfers_sheet(wb, config)

    # 3. Ward report sheets (9 wards)
    for ward in config.WARDS:
        build_ward_sheet(wb, config, ward)

    # 3b. Combined Emergency sheet (MAE + FAE side-by-side)
    build_emergency_combined_sheet(wb, config)

    # 4. Monthly Summary
    build_monthly_summary_sheet(wb, config)

    # 5. Ages Summary
    build_ages_summary_sheet(wb, config)

    # 6. Deaths Report
    build_deaths_report_sheet(wb, config)

    # 7. COD Summary
    build_cod_summary_sheet(wb, config)
    
    # 8. Statement of Inpatient
    build_statement_of_inpatient_sheet(wb, config)
    
    # 9. Non-Insured Report
    build_non_insured_report_sheet(wb, config)

    # Save
    wb.save(output_path)
    print(f"Phase 1 complete: {output_path}")
