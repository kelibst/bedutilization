import openpyxl
wb = openpyxl.Workbook()
ws = wb.active
ws['B2'] = 'All Wards'
ws['B3'] = 'All Months'
# test 2
ws['A1'] = '=COUNTIFS(A:A, IF($B$2="All Wards", "?*", $B$2), B:B, IF($B$3="All Months", ">=0", $B$3))'
wb.save('test_formulas.xlsx')

import win32com.client
excel = win32com.client.DispatchEx('Excel.Application')
excel.DisplayAlerts = False
wb2 = excel.Workbooks.Open(r'C:\Users\HIHMH\Desktop\projects\bedutilization\test_formulas.xlsx')
try:
    wb2.SaveAs(r'C:\Users\HIHMH\Desktop\projects\bedutilization\test_formulas.xlsm', FileFormat=52)
    print('Success')
except Exception as e:
    print('Failed:', e)
wb2.Close(SaveChanges=False)
excel.Quit()
